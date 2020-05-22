from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, Side
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.http.response import HttpResponse
from django.views.generic import TemplateView
from openpyxl.utils import get_column_letter
from django.urls import reverse_lazy
from django.db.models import Sum, Q
from openpyxl.styles import colors
from apps.index.models import *
from openpyxl.cell import Cell
from openpyxl import Workbook


def get_or_none(classmodel, **kwargs):
    try:
        return classmodel.objects.get(**kwargs)
    except classmodel.DoesNotExist:
        return None


def get_one_or_none(classmodel, **kwargs):
    query = classmodel.objects.filter(**kwargs).order_by('-id')
    if query.count() == 0:
        return None
    else:
        return query.first()


@method_decorator(login_required(login_url=reverse_lazy('index:login')), name='dispatch')
class ReportAircratfExcel(TemplateView):

    def get(self, request, *args, **kwargs):
        profile = request.user.profile
        aircratf_model = AirCraftModel.objects.all()
        crew = Crew.objects.all()
        wb = Workbook()
        
        start = request.GET.get('start_date',  None)
        end = request.GET.get('end_date',  None)

        for aircratf_model in aircratf_model:
            name = aircratf_model.name
            flightreport = FlightReport.objects.all().filter(
                aircraft__air_craft_models__name=name)
            
            if start is not None and end is not None:
                flightreport = flightreport.filter(date__range=[start, end])
            
            if profile.user_type.code == 3 or profile.user_type.code == 5:
                flightreport = flightreport.filter(aviation_unit=profile.tactic_unit)

            ws = wb.active
            ws = wb.create_sheet(name)
            ws.title = name

            ws.row_dimensions[1].height = 50

            redFill = PatternFill(start_color='ff0000',
                                  end_color='ff0000',
                                  fill_type='solid')

            bluefill = PatternFill(start_color='8db3e2',
                                   end_color='8db3e2',
                                   fill_type='solid')

            centered_alignment = Alignment(horizontal='center')
            border_bottom = Border(left=Side(border_style='medium',
                                             color='00000000'),
                                   right=Side(border_style='medium',
                                              color='00000000'),
                                   top=Side(border_style='medium',
                                            color='00000000'),
                                   bottom=Side(border_style='medium',
                                               color='00000000'),
                                   vertical=Side(border_style='medium',
                                                 color='00000000'),
                                   horizontal=Side(border_style='medium',
                                                   color='00000000'),
                                   )
            wrapped_alignment = Alignment(
                horizontal='general',
                vertical='center',
                wrap_text=True,
                shrink_to_fit=True
            )

            columns = [
                '    FECHA    ',
                '   HORA   ',
                '   UNIDAD DE AVIACION   ',
                'DIVISION/CONVENIO/FUERZA DE TAREA A QUIEN VAN CARGADAS LAS HORAS',
                '      UNIDAD OPERATIVA MENOR APOYADA      ',
                '   MATRICULA   ',
                '          TIPO MISIÓN          ',
                '   CONFIGURACION  ',
                '   CLASIFICACION DEL RIESGO   ',
                '   HRS MAQUINA   ',
                '   HRS TRIPULACION   ',
                '   OPERACIÓN MAYOR   ',
                '   NOMBRE DE LA OPERACION   ',
                '   DEPARTAMENTO   ',
                '   MUNICIPIO   ',
                '   RUTA   ',
                '   PAX   ',
                'ENFERMOS PROPIAS TROPAS',
                'HERIDOS PROPIAS TROPAS',
                'MUERTOS PROPIAS TROPAS',
                '   ENFERMOS ENEMIGO   ',
                '   HERIDOS ENEMIGO   ',
                '   MUERTOS ENEMIGO   ',
                '   EVACUACION CIVILES   ',
                '   KILOS   ',
                '   COMBUSTIBLE   ',
                '   PAM   ',
                '    ',
                '   P   ',
                '   ',
                '   IV  ',
                '   ',
                '   JT   ',
                '    ',
                '   TV   ',
                '    ',
                '   ART   ',
                '   ',
                '   CMA   ',
                '   ',
                '   OMI  ',
                '   ',
                '   OVE  ',
                '   ',
                '   ORDEN DE VUELO   ',
                '   EVENTO DE AVIACIÓN  ',
                '   OBSERVACIONES  ',
            ]

            column_widths = {}
            cont = 1
            for col_num, column_title in enumerate(columns, 1):
                cell = ws.cell(row=cont, column=col_num)
                cell.value = column_title
                cell.border = border_bottom
                cell.alignment = centered_alignment
                cell.alignment = wrapped_alignment
                cell.fill = bluefill
                column_letter = get_column_letter(col_num)
                if column_letter not in column_widths:
                    column_widths[column_letter] = 0
                if len(str(column_title)) >= column_widths[column_letter]:
                    column_widths[column_letter] = len(str(column_title)) + 1

            ws['A1'].fill = redFill
            ws['C1'].fill = redFill
            ws['D1'].fill = redFill
            ws['G1'].fill = redFill
            ws['H1'].fill = redFill

            ws.merge_cells('AE1:AF1')
            ws.merge_cells('AG1:AH1')
            ws.merge_cells('AI1:AJ1')
            ws.merge_cells('AK1:AL1')
            ws.merge_cells('AM1:AN1')
            ws.merge_cells('AO1:AP1')
            ws.merge_cells('AQ1:AR1')

            cont = 2
            for flightReport in flightreport:
                cell = ws.cell(row=cont, column=1)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=1).value = flightReport.date

                cell = ws.cell(row=cont, column=2)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=2).value = flightReport.time

                cell = ws.cell(row=cont, column=3)
                cell.alignment = centered_alignment
                ws.cell(
                    row=cont, column=3).value = flightReport.aviation_unit.abbreviation

                cell = ws.cell(row=cont, column=4)
                cell.alignment = centered_alignment
                ws.cell(
                    row=cont, column=4).value = flightReport.supported_unit.abbreviation

                cell = ws.cell(row=cont, column=5)
                cell.alignment = centered_alignment
                ws.cell(
                    row=cont, column=5).value = flightReport.charged_unit.abbreviation

                cell = ws.cell(row=cont, column=6)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=6).value = flightReport.aircraft.enrollment

                cell = ws.cell(row=cont, column=7)
                cell.alignment = centered_alignment
                ws.cell(
                    row=cont, column=7).value = flightReport.aviation_mission.mission_type.name

                cell = ws.cell(row=cont, column=8)
                cell.alignment = centered_alignment
                ws.cell(
                    row=cont, column=8).value = flightReport.configuration.abbreviation

                cell = ws.cell(row=cont, column=9)
                cell.alignment = centered_alignment
                ws.cell(
                    row=cont, column=9).value = flightReport.get_risk_classification_display()

                cell = ws.cell(row=cont, column=10)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=10).value = flightReport.machine_hours

                cell = ws.cell(row=cont, column=11)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=11).value = flightReport.crew_hours

                cell = ws.cell(row=cont, column=12)
                cell.alignment = centered_alignment
                ws.cell(
                    row=cont, column=12).value = flightReport.operation.major_operations.name

                cell = ws.cell(row=cont, column=13)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=13).value = flightReport.operation.name

                cell = ws.cell(row=cont, column=14)
                cell.alignment = centered_alignment
                ws.cell(
                    row=cont, column=14).value = flightReport.municipality.department.name

                cell = ws.cell(row=cont, column=15)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=15).value = flightReport.municipality.name

                cell = ws.cell(row=cont, column=16)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=16).value = flightReport.route

                cell = ws.cell(row=cont, column=17)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=17).value = flightReport.pax

                cell = ws.cell(row=cont, column=18)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=18).value = flightReport.sick_pt

                cell = ws.cell(row=cont, column=19)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=19).value = flightReport.wounded_pt

                cell = ws.cell(row=cont, column=20)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=20).value = flightReport.dead_pt

                cell = ws.cell(row=cont, column=21)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=21).value = flightReport.sick_en

                cell = ws.cell(row=cont, column=22)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=22).value = flightReport.wounded_en

                cell = ws.cell(row=cont, column=23)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=23).value = flightReport.dead_en

                cell = ws.cell(row=cont, column=24)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=24).value = flightReport.civil_evacuations

                cell = ws.cell(row=cont, column=25)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=25).value = flightReport.kilos

                cell = ws.cell(row=cont, column=26)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=26).value = flightReport.fuel

                crew = flightReport.crew_flight_report.all()

                pam = crew.filter(crew__flight_charge=Crew.PAM)
                if pam.count() > 0:
                    pam = pam.first()
                    cell = ws.cell(row=cont, column=27)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=27).value = pam.crew.rank.abbreviation
                    cell = ws.cell(row=cont, column=28)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=28).value = pam.crew.name

                p = crew.filter(crew__flight_charge=Crew.PI)
                if p.count() > 0:
                    p = p.first()
                    cell = ws.cell(row=cont, column=29)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=29).value = p.crew.rank.abbreviation
                    cell = ws.cell(row=cont, column=30)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=30).value = p.crew.name

                iv = crew.filter(crew__flight_charge=Crew.IV)
                if iv.count() > 0:
                    iv = iv.first()
                    cell = ws.cell(row=cont, column=31)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=31).value = iv.crew.rank.abbreviation
                    cell = ws.cell(row=cont, column=32)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=32).value = iv.crew.name

                jt = crew.filter(crew__flight_charge=Crew.JT)
                if jt.count() > 0:
                    jt = jt.first()
                    cell = ws.cell(row=cont, column=33)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=33).value = jt.crew.rank.abbreviation
                    cell = ws.cell(row=cont, column=34)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=34).value = jt.crew.name

                tv = crew.filter(crew__flight_charge=Crew.TV)
                if tv.count() > 0:
                    tv = tv.first()
                    cell = ws.cell(row=cont, column=35)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=35).value = tv.crew.rank.abbreviation
                    cell = ws.cell(row=cont, column=36)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=36).value = tv.crew.name

                art = crew.filter(crew__flight_charge=Crew.ART)
                if art.count() > 0:
                    art = art.first()
                    cell = ws.cell(row=cont, column=37)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=37).value = art.crew.rank.abbreviation
                    cell = ws.cell(row=cont, column=38)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=38).value = art.crew.name

                cma = crew.filter(crew__flight_charge=Crew.CMA)
                if cma.count() > 0:
                    cma = cma.first()
                    cell = ws.cell(row=cont, column=39)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=39).value = cma.crew.rank.abbreviation
                    cell = ws.cell(row=cont, column=40)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=40).value = cma.crew.name

                omi = crew.filter(crew__flight_charge=Crew.OMI)
                if omi.count() > 0:
                    omi = omi.first()
                    cell = ws.cell(row=cont, column=41)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=41).value = omi.crew.rank.abbreviation
                    cell = ws.cell(row=cont, column=42)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=42).value = omi.crew.name

                ove = crew.filter(crew__flight_charge=Crew.OVE)
                if ove.count() > 0:
                    ove = ove.first()
                    cell = ws.cell(row=cont, column=43)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=43).value = ove.crew.rank.abbreviation
                    cell = ws.cell(row=cont, column=44)
                    cell.alignment = centered_alignment
                    ws.cell(row=cont, column=44).value = ove.crew.name

                cell = ws.cell(row=cont, column=45)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=45).value = flightReport.flight_order_id

                cell = ws.cell(row=cont, column=46)
                cell.alignment = centered_alignment
                aviation_event = '---'
                if flightReport.aviation_event is not None:
                    aviation_event = flightReport.aviation_event.name
                ws.cell(row=cont, column=46).value = aviation_event

                cell = ws.cell(row=cont, column=47)
                cell.alignment = centered_alignment
                ws.cell(row=cont, column=47).value = flightReport.observations

                for col_num, column_title in enumerate(columns, 1):
                    cell = ws.cell(row=cont, column=col_num)
                    column_letter = get_column_letter(col_num)
                    if column_letter not in column_widths:
                        column_widths[column_letter] = 0
                    if len(str(cell.value)) >= column_widths[column_letter]:
                        column_widths[column_letter] = len(str(cell.value)) + 1

                cont = cont + 1

            for col_num, column_title in enumerate(columns, 1):
                column_letter = get_column_letter(col_num)
                column_dimensions = ws.column_dimensions[column_letter]
                column_dimensions.width = column_widths[column_letter]

        ws.merge_cells(
            start_row=cont+1, start_column=1, end_row=cont+2, end_column=len(columns))
        cell = ws.cell(row=cont+1, column=1)
        #cell.value = footer_usuarios
        cell.alignment = wrapped_alignment
        del wb['Sheet']
        filename = "ReporteDeAeronavesExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename={0}".format(filename)
        response["Content-Disposition"] = content

        wb.save(response)
        return response


def get_assisted_unit(data):
    result = {}
    is_agreement = True
    if data['agreement'] is not None:
        result['assisted_unit'] = data['agreement']
        result['assisted_unit__name'] = data['agreement__agreement_name']
        result['assisted_unit__abbreviation'] = data['agreement__abbreviation']
    elif data['major_operative_unit'] is not None:
        result['assisted_unit'] = data['major_operative_unit']
        result['assisted_unit__name'] = data['major_operative_unit__name']
        result['assisted_unit__abbreviation'] = data['major_operative_unit__abbreviation']
        is_agreement = False
    result['is_agreement'] = is_agreement
    return result


def get_default_hours(aircraft_models, assisted_unit_pk, is_agreement):
    default_hours_model = {}
    for aircraft_model in aircraft_models:
        hours_result = {}
        assigned_hours = 0
        if assisted_unit_pk is not None:
            if is_agreement:
                assigned = get_one_or_none(
                    AssignedHourAgreementAircraftModel, agreement__pk=assisted_unit_pk, aircraft_model=aircraft_model)
            else:
                assigned = get_one_or_none(AssignedHourMajorOperationAircraftModel,
                                           major_operative_unit__pk=assisted_unit_pk, aircraft_model=aircraft_model)
            if assigned is not None:
                assigned_hours = assigned.assigned_hours
        hours_result['assigned_hours__sum'] = assigned_hours
        hours_result['fly_hours__sum'] = 0
        hours_result['available_hours__sum'] = assigned_hours
        default_hours_model[aircraft_model.pk] = hours_result
    return default_hours_model


@method_decorator(login_required(login_url=reverse_lazy('index:login')), name='dispatch')
class ExcelHoursAircraft(TemplateView):

    def get(self, request, *args, **kwargs):
        profile = request.user.profile
        aircraft_models = AirCraftModel.objects.all().order_by('pk')
        aircraft_models_dict = {}
        aviation_units = TacticUnit.objects.filter(
            is_aviation=True).order_by('pk')
        count_aviation = {}
        real_query = {}
        total_hours = get_default_hours(aircraft_models, None, None)
        
        start = request.GET.get('start_date',  None)
        end = request.GET.get('end_date',  None)
        
        basic_query_flight_report = FlightReport.objects.all()
        if start is not None and end is not None:
            basic_query_flight_report = basic_query_flight_report.filter(date__range=[start, end])
        
        if profile.user_type.code == 3 or profile.user_type.code == 5:
            basic_query_flight_report = basic_query_flight_report.filter(aviation_unit=profile.tactic_unit)
        
        for aviation_unit in aviation_units:
            count_aviation[aviation_unit.pk] = 0
        for aircraft_model in aircraft_models:
            flight_reports = basic_query_flight_report.filter(aircraft__air_craft_models=aircraft_model).exclude(Q(aviation_mission__abbreviation='RA') | Q(
                aviation_mission__abbreviation='E') | Q(aviation_mission__abbreviation='M') | Q(aviation_mission__abbreviation='AEG')).select_related().order_by('major_operative_unit')
            queryset = flight_reports.values('aviation_unit', 'aviation_unit__name', 'aviation_unit__abbreviation', 'agreement', 'agreement__abbreviation', 'agreement__agreement_name',
                                             'major_operative_unit', 'major_operative_unit__name', 'major_operative_unit__abbreviation').order_by('aviation_unit').annotate(machine_hours__sum=Sum('machine_hours'))
            list_result = []
            # print(queryset.count())
            results_aviation = {}
            count_aviation = {}
            for aviation_unit in aviation_units:
                results_aviation[aviation_unit.pk] = []
                count_aviation[aviation_unit.pk] = 0
            for data in queryset:
                result = get_assisted_unit(data)
                result['aviation_unit'] = data['aviation_unit']
                result['aviation_unit__abbreviation'] = data['aviation_unit__abbreviation']
                result['aviation_unit__name'] = data['aviation_unit__name']
                assigned_hours = 0
                if result['is_agreement']:
                    assigned = get_one_or_none(
                        AssignedHourAgreementAircraftModel, agreement__pk=result['assisted_unit'], aircraft_model=aircraft_model)
                else:
                    assigned = get_one_or_none(AssignedHourMajorOperationAircraftModel,
                                               major_operative_unit__pk=result['assisted_unit'], aircraft_model=aircraft_model)
                if assigned is not None:
                    assigned_hours = assigned.assigned_hours
                hours_result = {}
                hours_result['assigned_hours__sum'] = assigned_hours
                hours_result['fly_hours__sum'] = data['machine_hours__sum']
                hours_result['available_hours__sum'] = hours_result['assigned_hours__sum'] - \
                    hours_result['fly_hours__sum']
                key_query = str(result['aviation_unit']) + \
                    '_' + result['assisted_unit__abbreviation']
                if key_query in real_query:
                    real_query[key_query]['aicraft_models'][aircraft_model.pk]['assigned_hours__sum'] += hours_result[
                        'assigned_hours__sum']
                    real_query[key_query]['aicraft_models'][aircraft_model.pk]['fly_hours__sum'] += hours_result[
                        'fly_hours__sum']
                    real_query[key_query]['aicraft_models'][aircraft_model.pk]['available_hours__sum'] += hours_result[
                        'available_hours__sum']
                else:
                    result['aicraft_models'] = get_default_hours(
                        aircraft_models, result['assisted_unit'], result['is_agreement'])
                    result['aicraft_models'][aircraft_model.pk] = hours_result
                    real_query[key_query] = result
        for key_query, row_table in real_query.items():
            for aircraft_model_pk, hours in row_table['aicraft_models'].items():
                total_hours[aircraft_model_pk]['assigned_hours__sum'] += real_query[key_query]['aicraft_models'][aircraft_model_pk]['assigned_hours__sum']
                total_hours[aircraft_model_pk]['fly_hours__sum'] += real_query[key_query]['aicraft_models'][aircraft_model_pk]['fly_hours__sum']
                total_hours[aircraft_model_pk]['available_hours__sum'] += real_query[key_query]['aicraft_models'][aircraft_model_pk]['available_hours__sum']
        # full_aviation_units = []
        # for aviation_unit in aviation_units:
        #     for count in range(count_aviation[aviation_unit.pk]+1):
        #         full_aviation_units.append(aviation_unit)
        real_query = dict(sorted(real_query.items()))
        wb = Workbook()
        #name = aircratf.name
        ws = wb.active
        ws.title = "Horas Unidades"
        ws.row_dimensions[1].height = 50
        centered_alignment = Alignment(horizontal='center')
        border_bottom = Border(left=Side(border_style='medium',
                                         color='00000000'),
                               right=Side(border_style='medium',
                                          color='00000000'),
                               top=Side(border_style='medium',
                                        color='00000000'),
                               bottom=Side(border_style='medium',
                                           color='00000000'),
                               vertical=Side(border_style='medium',
                                             color='00000000'),
                               horizontal=Side(border_style='medium',
                                               color='00000000'),
                               )
        wrapped_alignment = Alignment(
            horizontal='general',
            vertical='center',
            wrap_text=True,
            shrink_to_fit=True
        )
        redFill = PatternFill(start_color='ff0000',
                              end_color='ff0000',
                              fill_type='solid')

        bluefill = PatternFill(start_color='8db3e2',
                               end_color='8db3e2',
                               fill_type='solid')

        greenfill = PatternFill(start_color='228b22',
                                end_color='228b22',
                                fill_type='solid')

        red2fill = PatternFill(start_color='f13131',
                               end_color='f13131',
                               fill_type='solid')

        yellowfill = PatternFill(start_color='faf32f',
                                 end_color='faf32f',
                                 fill_type='solid')

        columns = [
            '    UNIDAD DE AVIACIÓN ',
            '    UNIDAD APOYADA',
        ]

        second_columns = [
            '    UNIDAD DE AVIACIÓN ',
            '    UNIDAD APOYADA',
        ]

        columns_count = 3

        for aircraft_model in aircraft_models:
            columns.append(aircraft_model.name)
            columns.append('')
            columns.append('')
            # Second columns
            second_columns.append('ASIG')
            second_columns.append('VOL')
            second_columns.append('DISP')

        column_widths = {}
        cont = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=cont, column=col_num)
            cell.value = column_title
            cell.border = border_bottom
            cell.alignment = centered_alignment
            cell.alignment = wrapped_alignment
            cell.fill = bluefill
            column_letter = get_column_letter(col_num)
            if column_letter not in column_widths:
                column_widths[column_letter] = 0
            if len(str(column_title)) >= column_widths[column_letter]:
                column_widths[column_letter] = len(str(column_title)) + 5
            # Second columns
            cell = ws.cell(row=cont+1, column=col_num)
            cell.value = second_columns[col_num-1]
            cell.border = border_bottom
            cell.alignment = centered_alignment
            cell.alignment = wrapped_alignment
            cell.fill = bluefill
            column_letter = get_column_letter(col_num)
            if column_letter not in column_widths:
                column_widths[column_letter] = 0
            if len(str(second_columns[col_num-1])) >= column_widths[column_letter]:
                column_widths[column_letter] = len(
                    str(second_columns[col_num-1])) + 5
        cont += 1
        # ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        # ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        columns_count = 2
        for aircraft_model in aircraft_models:
            columns_count += 1
            ws.merge_cells(start_row=1, start_column=columns_count,
                           end_row=1, end_column=columns_count+2)
            columns_count += 2

        # Table Rows
        for key, row_table in real_query.items():
            cont += 1
            row = [
                row_table['aviation_unit__abbreviation'],
                row_table['assisted_unit__abbreviation'],
            ]

            for air_key, aircraft_model in row_table['aicraft_models'].items():
                row.append(aircraft_model['assigned_hours__sum'])
                row.append(aircraft_model['fly_hours__sum'])
                row.append(aircraft_model['available_hours__sum'])

            for col_num, column_title in enumerate(row, 1):
                cell = ws.cell(row=cont, column=col_num)
                cell.value = column_title
                cell.alignment = wrapped_alignment
                if col_num <= 2:
                    cell.fill = bluefill
                cell.border = border_bottom
                column_letter = get_column_letter(col_num)
                if column_letter not in column_widths:
                    column_widths[column_letter] = 0
                if len(str(column_title)) >= column_widths[column_letter]:
                    column_widths[column_letter] = len(str(column_title)) + 1

        cont += 1
        # Total Row
        row = [
            '  TOTAL  ',
            '  TOTAL  ',
        ]

        for total_key, total in total_hours.items():
            row.append(total['assigned_hours__sum'])
            row.append(total['fly_hours__sum'])
            row.append(total['available_hours__sum'])

        for col_num, column_title in enumerate(row, 1):
            cell = ws.cell(row=cont, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.fill = bluefill
            cell.border = border_bottom
            column_letter = get_column_letter(col_num)
            if column_letter not in column_widths:
                column_widths[column_letter] = 0
            if len(str(column_title)) >= column_widths[column_letter]:
                column_widths[column_letter] = len(str(column_title)) + 5

        ws.merge_cells(start_row=cont, start_column=1,
                       end_row=cont, end_column=2)

        # merge columns 1 and 2
        for col_num in range(2):
            col_num += 1
            cell = ws.cell(row=1, column=col_num)
            cell_value = cell.value
            cell_row = 1
            for row_count in range(len(real_query)+1):
                row_count += 2
                cell = ws.cell(row=row_count, column=col_num)
                current_cell_value = cell.value
                if current_cell_value != cell_value:
                    cell_value = current_cell_value
                    ws.merge_cells(start_row=cell_row, start_column=col_num,
                                   end_row=row_count-1, end_column=col_num)
                    cell_row = row_count
            ws.merge_cells(start_row=cell_row, start_column=col_num,
                           end_row=row_count, end_column=col_num)

        # Second table

        missions = [{'mission': 'RA', 'search': True, 'col_1': 'DAVAA', 'col_2': 'RA (BRIAV33)'}, {'mission': 'E', 'search': True, 'col_1': 'DAVAA', 'col_2': 'E (BRIAV33)'}, {'mission': 'E', 'search': False, 'col_1': 'DAVAA', 'col_2': 'E (BRIAV25)'}, {
            'mission': 'M', 'search': True, 'col_1': 'DAVAA', 'col_2': 'M (EN AREA)'}, {'mission': 'AEG', 'search': True, 'col_1': 'DAVAA', 'col_2': 'AEG'}]

        list_row = []
        total_missions = get_default_hours(aircraft_models, None, None)

        for mission in missions:
            row = []
            row.append(mission['col_1'])
            row.append(mission['col_2'])
            for aircraft_model in aircraft_models:
                flight_reports = basic_query_flight_report.filter(
                    aircraft__air_craft_models=aircraft_model, aviation_mission__abbreviation=mission['mission']).select_related().order_by('major_operative_unit')
                if not mission['search']:
                    flight_reports = FlightReport.objects.none()
                queryset = flight_reports.values('aviation_unit', 'aviation_unit__name', 'aviation_unit__abbreviation', 'agreement', 'agreement__abbreviation', 'agreement__agreement_name',
                                                 'major_operative_unit', 'major_operative_unit__name', 'major_operative_unit__abbreviation').order_by('aviation_unit').annotate(machine_hours__sum=Sum('machine_hours'))
                result = {'fly_hours__sum': 0}
                for data in queryset:
                    result['fly_hours__sum'] += data['machine_hours__sum']
                row.append('')
                row.append(result['fly_hours__sum'])
                row.append('')
                total_missions[aircraft_model_pk]['fly_hours__sum'] += result['fly_hours__sum']
            list_row.append(row)

        cont += 1

        for row in list_row:
            cont += 1
            for col_num, column_title in enumerate(row, 1):
                cell = ws.cell(row=cont, column=col_num)
                cell.value = column_title
                cell.alignment = wrapped_alignment
                if col_num <= 2:
                    cell.fill = bluefill
                cell.border = border_bottom
                column_letter = get_column_letter(col_num)
                if column_letter not in column_widths:
                    column_widths[column_letter] = 0
                if len(str(column_title)) >= column_widths[column_letter]:
                    column_widths[column_letter] = len(str(column_title)) + 1

        row = [
            'TOTAL VOLADO BRIAV33',
            'TOTAL VOLADO BRIAV33',
        ]
        aircraft_models = list(aircraft_models)
        for array_pos in range(len(aircraft_models)):
            aircraft_model = aircraft_models[array_pos]
            fly_hours = total_missions[aircraft_model.pk]['fly_hours__sum'] + \
                total_hours[aircraft_model.pk]['fly_hours__sum']
            row.append(fly_hours)
            row.append(fly_hours)
            row.append(fly_hours)

        cont += 1
        current_cont = cont
        for col_num, column_title in enumerate(row, 1):
            cell = ws.cell(row=cont, column=col_num)
            cell.value = column_title
            cell.alignment = centered_alignment
            if col_num <= 2:
                cell.fill = bluefill
            cell.border = border_bottom
            column_letter = get_column_letter(col_num)
            if column_letter not in column_widths:
                column_widths[column_letter] = 0
            if len(str(column_title)) >= column_widths[column_letter]:
                column_widths[column_letter] = len(str(column_title)) + 5

        value_first_column = 0
        cell = ws.cell(row=cont, column=3)
        value_first_column += cell.value
        cell = ws.cell(row=cont, column=6)
        value_first_column += cell.value
        for idx in range(3, 9):
            cell = ws.cell(row=cont, column=idx)
            cell.value = value_first_column

        # AVALIABLE ROW
        cont += 1
        row = [
            'DISPONIBILIDAD DE AERONAVES',
            'DISPONIBILIDAD DE AERONAVES',
        ]

        for aircraft_model in aircraft_models:
            flight_reports = basic_query_flight_report.filter(aircraft__air_craft_models=aircraft_model).select_related().distinct(
                'aviation_unit', 'major_operative_unit', 'agreement', 'aircraft').order_by('aircraft', 'major_operative_unit')
            sum_avaliable_hours = 0
            for flight_report in flight_reports:
                sum_avaliable_hours += flight_report.aircraft.hours_available
            row.append(sum_avaliable_hours)
            row.append(sum_avaliable_hours)
            row.append(sum_avaliable_hours)

        for col_num, column_title in enumerate(row, 1):
            cell = ws.cell(row=cont, column=col_num)
            cell.value = column_title
            cell.alignment = centered_alignment
            if col_num <= 2:
                cell.fill = bluefill
            cell.border = border_bottom
            column_letter = get_column_letter(col_num)
            if column_letter not in column_widths:
                column_widths[column_letter] = 0
            if len(str(column_title)) >= column_widths[column_letter]:
                column_widths[column_letter] = len(str(column_title)) + 5

        for idx_merge in range(2):
            cont_merge = current_cont + idx_merge
            ws.merge_cells(start_row=cont_merge, start_column=1,
                           end_row=cont_merge, end_column=2)

            ws.merge_cells(start_row=cont_merge, start_column=3,
                           end_row=cont_merge, end_column=8)

            col_count = 2
            start_column = col_count
            for array_pos in range(len(aircraft_models)):
                start_column += 1
                end_column = start_column + 2
                if array_pos > 1:
                    ws.merge_cells(start_row=cont_merge, start_column=start_column,
                                   end_row=cont_merge, end_column=end_column)
                start_column = end_column

        # TABLE 3 AIRCRAFT:
        aircraft_states = [{'state': 'AMI', 'fill': red2fill}, {'state': 'AMP', 'fill': red2fill}, {
            'state': 'AVP', 'fill': red2fill}, {'state': 'LRM', 'fill': yellowfill}, {'state': 'ACL', 'fill': greenfill}]
        row = [
            '',
        ]
        row.append('A/A')
        for state in aircraft_states:
            row.append(state['state'])
        row.append('Operacionales')

        cont += 2
        table_4_cont = cont

        for col_num, column_title in enumerate(row, 1):
            cell = ws.cell(row=cont, column=col_num)
            cell.value = column_title
            cell.alignment = centered_alignment
            cell.border = border_bottom
            column_letter = get_column_letter(col_num)
            if column_letter not in column_widths:
                column_widths[column_letter] = 0
            if len(str(column_title)) >= column_widths[column_letter]:
                column_widths[column_letter] = len(str(column_title)) + 5

        for state_idx in range(len(aircraft_states)):
            cell = ws.cell(row=cont, column=state_idx+3)
            cell.fill = aircraft_states[state_idx]['fill']

        aircrafts_dicts = {}
        total_states = [0, 0, 0, 0, 0, 0, 0]
        for aircraft_model in aircraft_models:
            aircraft_query = AirCraft.objects.filter(
                is_active=True, air_craft_models=aircraft_model)
            a_a = aircraft_query.count()
            total_states[0] += a_a
            states_list = []
            op = 0
            for state_idx in range(len(aircraft_states)):
                state_count = aircraft_query.filter(
                    aircraft_status=aircraft_states[state_idx]['state']).count()
                states_list.append(state_count)
                total_states[state_idx+1] += state_count
                if aircraft_states[state_idx]['state'] == 'ACL' or aircraft_states[state_idx]['state'] == 'LRM':
                    op += state_count
            total_states[6] += op
            aircrafts_dicts[aircraft_model.pk] = {
                'a_a': a_a, 'states': states_list, 'op': op}

        list_row = [['UH-60 / S-70', 0, 0, 0, 0, 0, 0, 0]]

        for air_idx in range(len(aircraft_models)):
            if air_idx < 2:
                current_row = list_row[0]
                current_row[1] += aircrafts_dicts[aircraft_models[air_idx].pk]['a_a']
                cont_temp = 2
                states = aircrafts_dicts[aircraft_models[air_idx].pk]['states']
                for state in states:
                    current_row[cont_temp] += state
                    cont_temp += 1
                current_row[cont_temp] += aircrafts_dicts[aircraft_models[air_idx].pk]['op']
                list_row[0] = current_row
            else:
                row = [aircraft_models[air_idx].name]
                row.append(aircrafts_dicts[aircraft_models[air_idx].pk]['a_a'])
                states = aircrafts_dicts[aircraft_models[air_idx].pk]['states']
                for state in states:
                    row.append(state)
                row.append(aircrafts_dicts[aircraft_models[air_idx].pk]['op'])
                list_row.append(row)
        list_row.append(['TOTAL', *total_states])

        for row in list_row:
            cont += 1
            for col_num, column_title in enumerate(row, 1):
                cell = ws.cell(row=cont, column=col_num)
                cell.value = column_title
                cell.alignment = wrapped_alignment
                cell.border = border_bottom
                column_letter = get_column_letter(col_num)
                if column_letter not in column_widths:
                    column_widths[column_letter] = 0
                if len(str(column_title)) >= column_widths[column_letter]:
                    column_widths[column_letter] = len(str(column_title)) + 5
        
        cont += 1
        cell = ws.cell(row=cont, column=2)
        cell.value = '% ALISTAMIENTO'
        cell.border = border_bottom

        cell = ws.cell(row=cont, column=3)
        cell.value = (total_states[-1] / total_states[0]) / 100
        cell.number_format = '0.00%'
        cell.border = border_bottom
        
        ws.merge_cells(start_row=cont, start_column=3, end_row=cont, end_column=8)

        # TABLE 4
        cont = table_4_cont - 1
        column = 10

        list_row = []

        row = ['EQUIPO', 'ASIG', 'DISP', '% VOLADO']
        list_row.append(row)

        total = ['TOTAL', 0, 0, 0]

        for aircraft_model in aircraft_models:
            row = []
            row.append(aircraft_model.name)
            assigned_hours = total_hours[aircraft_model.pk]['assigned_hours__sum']
            row.append(assigned_hours)
            total[1] += assigned_hours
            row.append(total_hours[aircraft_model.pk]['available_hours__sum'])
            total[2] += total_hours[aircraft_model.pk]['available_hours__sum']
            fly_hours = total_missions[aircraft_model.pk]['fly_hours__sum'] + \
                total_hours[aircraft_model.pk]['fly_hours__sum']
            percentage = 0
            if assigned_hours > 0:
                percentage = float(fly_hours/assigned_hours)
            row.append(percentage)
            total[3] += percentage
            list_row.append(row)

        list_row.append(total)

        for row in list_row:
            cont += 1
            for col_num, column_title in enumerate(row, 1):
                cell = ws.cell(row=cont, column=col_num+column)
                cell.value = column_title
                cell.alignment = wrapped_alignment
                cell.border = border_bottom
                if col_num >= 4:
                    cell.number_format = '0.00%'
                column_letter = get_column_letter(col_num)
                if column_letter not in column_widths:
                    column_widths[column_letter] = 0
                if len(str(column_title)) + 5 >= column_widths[column_letter]:
                    column_widths[column_letter] = len(str(column_title)) + 5
        
        
        # TABLE 5
        cont = table_4_cont - 1
        column = 15
        
        list_row = [
            [{'content': 'CONVENCIONES', 'fill': None}, {'content': 'CONVENCIONES', 'fill': None}],
            [{'content': '', 'fill': greenfill}, {'content': '10 O MAS HORAS', 'fill': None}],
            [{'content': '', 'fill': yellowfill}, {'content': '5 A 9 HORAS', 'fill': None}],
            [{'content': '', 'fill': red2fill}, {'content': '0 a 4 HORAS', 'fill': None}],
            [{'content': 'EJC XXXX', 'fill': None}, {'content': 'ASEGURADAS', 'fill': None}],
            [{'content': 'ACL', 'fill': greenfill}, {'content': 'ACL', 'fill': None}],
            [{'content': 'LRM', 'fill': yellowfill}, {'content': 'LRM', 'fill': None}],
            [{'content': 'XXX', 'fill': red2fill}, {'content': 'AMI, AMP, AVP,APA', 'fill': None}],
        ]
        
        for row in list_row:
            cont += 1
            for col_num, column_title in enumerate(row, 1):
                cell = ws.cell(row=cont, column=col_num+column)
                cell.value = column_title['content']
                cell.alignment = wrapped_alignment
                cell.border = border_bottom
                if column_title['fill'] is not None:
                    cell.fill = column_title['fill']
                column_letter = get_column_letter(col_num)
                if column_letter not in column_widths:
                    column_widths[column_letter] = 0
                if len(str(column_title['content'])) >= column_widths[column_letter]:
                    column_widths[column_letter] = len(str(column_title['content'])) + 5
        
        ws.merge_cells(start_row=table_4_cont, start_column=16, end_row=table_4_cont, end_column=17)
        
        # for col_num, column_title in enumerate(second_columns, 1):
        #     column_letter = get_column_letter(col_num)
        #     column_dimensions = ws.column_dimensions[column_letter]
        #     column_dimensions.width = column_widths[column_letter]

        filename = "ReporteDeUnidadesExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename={0}".format(filename)
        response["Content-Disposition"] = content

        wb.save(response)
        return response
